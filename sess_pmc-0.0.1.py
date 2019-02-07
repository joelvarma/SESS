
# coding: utf-8

# In[13]:


import xml.etree.cElementTree as ET
import openpyxl 

__author__ = 'Joelvarma'

wb = openpyxl.Workbook()
sheet = wb.active

dup_dict={}

file_path = "pmc_result.xml"

context = ET.iterparse(file_path, events=("start", "end"))

context = iter(context)

counter=r_num=r_num2=1

for event, elem in context:

    tag = elem.tag
    value = elem.text
    
    if counter > 1:
        counter+=1
    
    if event == 'start' :

        if tag == 'surname':
            
            tempsname=value
            
        if tag == 'given-names':
            
            tempgname=value
            counter = 0
            
        if tag == 'email' and counter <= 3 and counter > 0 and (prev_tag=='given-names' or prev_tag == 'address' or prev_tag=='phone'):
            
            #print("email",tempsname,tempgname,value)

            dup_dict[value]=1
            
            c1 = sheet.cell(row = r_num, column = 1)
            
            if tempsname == None:
                tempsname = ''

            if tempgname == None:
                tempgname = ''
                
            c1.value = str(tempsname) +" "+ str(tempgname)
            
            c2 = sheet.cell(row= r_num , column = 2) 
            c2.value = value
            r_num+=1
            
            counter = 0
                
    if event == 'end':
        
        if tag == 'given-names':
            counter+=1
            
    prev_tag = tag
    
    elem.clear()

wb.save("names_emails_pmc.xlsx")

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

prev_tag = None

context = ET.iterparse(file_path, events=("start", "end"))

context = iter(context)

dup_dict2={}
count=count2=0

for event,elem in context:

    tag = elem.tag
    value = elem.text

    if event == 'start':
        
        
        if tag == 'email'  and (prev_tag!='given-names' and prev_tag!='address' and prev_tag!='phone'):

            if dup_dict.get(value,None) == None:
                
                if dup_dict2.get(value,None) == None:
                
                    #print("email2",value)

                    dup_dict2[value]=1

                    c3 = sheet2.cell(row= r_num2 , column = 1) 

                    c3.value = value

                    r_num2+=1
                else:
                    count2+=1

            else:
                count+=1

    prev_tag = tag
    
    elem.clear()

wb2.save("emails_pmc.xlsx")

print("No. of Accurate emails found : ", r_num)

print("No. of Unknown emails found : ", r_num2)

print("No. of Duplicate emails eliminated : ", count,"+",count2,"=",count+count2)
print(r_num+r_num2+count+count2)

